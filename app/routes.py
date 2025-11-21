import os
import sys
import subprocess
from tempfile import NamedTemporaryFile

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from werkzeug.utils import secure_filename

# from app.web_run import realTimeProgress, realTimeText
from flask import (
    Blueprint,
    request,
    render_template,
    jsonify,
    send_file,
)

sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))
from scripts.navitime.helper import get_best_routes, get_commute_pass
from scripts.common.main import get_driver, is_on_application

sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))
from scripts.e_tax import main as script_main
from scripts.excel_compare import main as script_excel_compare

# Create a blueprint for the routes
main = Blueprint("main", __name__)


@main.route("/", methods=["GET", "POST"])
def index():
    if request.method == "POST":
        # Get the uploaded file from the request
        file = request.files["file"]
        tab = request.form["tab"]
        # Check if the file is an Excel file
        if file and file.filename.endswith(".xlsx"):
            # Read the Excel file into a DataFrame
            df = pd.read_excel(file)
            # Config
            config = get_config()
            # Create a new DataFrame to store results
            results = []
            # Create a new WebDriver instance
            driver = get_driver()
            if tab == "1":
                results = process_tab_1(df, driver, config)
            elif tab == "2":
                results = process_tab_2(df, driver)

            # Save results to an Excel file
            output_file = save_results_to_excel(results)

            # Close and quit the WebDriver
            if driver:
                driver.close()
                driver.quit()

            # Send the file to the client
            return send_file(
                output_file, as_attachment=True, download_name="Results.xlsx"
            )
    return render_template("app.html")


def get_config():
    return {
        "train": {"feature": "Feature (安、早、楽)", "money": "金額"},
        "car": {
            "feature": "Feature (有料、無料、距離、低燃費、景観)",
            "money": -1,
        },
        "bus": {"feature": "Feature (安、早、楽)", "money": "金額"},
        "walk": {"feature": "", "money": -1},
        "bike": {
            "feature": "Feature (距離が短い、坂道が少ない、坂道が多い、大通り優先、裏通り優先)",
            "money": -1,
        },
        "plane": {
            "feature": "航空会社 (すべて表示、ANAのみ、JALのみ、LCC/その他)",
            "money": -1,
        },
        "truck": {
            "feature": "Feature (推奨ルート、無料優先、高速優先、道幅優先)",
            "money": -1,
        },
    }


def process_tab_1(df, driver, config):
    results = []
    transport = request.form["transport"]
    for index_inner, row in df.iterrows():
        driver.get("https://www.navitime.co.jp/")
        id_ = row["ID"]
        name = row["名前"]
        start = row["出発点"]
        end = row["行き先"]
        feature_row, money = get_feature_and_money(row, config, transport)
        standard = row[feature_row] if len(feature_row) > 0 else ""
        data = get_best_routes(
            driver, start, end, money, standard, index_inner, len(df), transport
        )
        results.append(
            create_result(
                data, id_, name, start, end, money, standard, transport, feature_row
            )
        )
    return results


def get_feature_and_money(row, config, transport):
    if transport in config:
        feature_row = config[transport]["feature"]
        money = (
            row[config[transport]["money"]] if config[transport]["money"] != -1 else -1
        )
    else:
        feature_row = ""
        money = -1
    return feature_row, money


def create_result(data, id_, name, start, end, money, standard, transport, feature_row):
    result = {
        "ID": id_,
        "名前": name,
        "出発点": start,
        "行き先": end,
        "金額": money if money and money >= 0 else "",
        "入力されたFeature": standard if standard and len(standard) > 0 else "",
    }
    if transport == "plane":
        flight_keys_labels = {
            "flightName": "便名",
            "flightStartTime": "出発",
            "flightEndTime": "到着",
            "time": "時間",
        }

        for i in range(3):
            for flight_data, label in flight_keys_labels.items():
                if flight_data in data and len(data[flight_data]) > i:
                    result[f"{label} {i + 1}"] = data[flight_data][i]
    else:
        other_keys_labels = {
            "routes": "最適なルート",
            "cost": "金額",
            "feature": feature_row,
            "transferTime": "乗換回数",
            "time": "乗車時間・移動時間",
            "distance": "距離",
            "step": "歩数",
            "calories": "カロリー",
        }
        for key, label in other_keys_labels.items():
            if key in data:
                for i in range(min(3, len(data[key]))):
                    result[f"{label} {i + 1}"] = data[key][i]
    return result


def process_tab_2(df, driver):
    results = []
    for _, row in df.iterrows():
        driver.get("https://www.navitime.co.jp/transfer/pass/?fromlink=pcnavi.header")
        id_ = row["ID"]
        name = row["名前"]
        start = row["出発駅"]
        end = row["到着駅"]
        transitions = get_transitions(row)
        row_type = row["種別(通勤、通学、通学（高校）、通学（中学）)"]
        data = get_commute_pass(driver, start, end, transitions, row_type)
        results.append(
            create_commute_result(data, id_, name, start, end, transitions, row_type)
        )
    return results


def get_transitions(row):
    transition1 = row["経由駅1"]
    transition2 = row["経由駅2"]
    transition3 = row["経由駅3"]
    transitions = [transition1, transition2, transition3]
    transitions = [t if pd.notna(t) else "" for t in transitions]
    return transitions


def create_commute_result(data, id_, name, start, end, transitions, row_type):
    result = {
        "ID": id_,
        "名前": name,
        "出発点": start,
        "行き先": end,
        "経由駅１": transitions[0],
        "経由駅２": transitions[1],
        "経由駅３": transitions[2],
        "種別": row_type,
    }
    keys = ["routes", "price1month", "price3month", "price6month"]
    labels = ["ルート", "1ヶ月の価格", "3ヶ月の価格", "6ヶ月の価格"]
    for i in range(3):
        for key, label in zip(keys, labels):
            if key in data and len(data[key]) > i:
                result[f"{label} {i + 1}"] = data[key][i]
    return result


def save_results_to_excel(results):
    results_df = pd.DataFrame(results)
    with NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_file:
        output_file = temp_file.name
    with pd.ExcelWriter(output_file, engine="xlsxwriter") as writer:
        try:
            results_df.to_excel(writer, index=False, sheet_name="Sheet1")
            workbook = writer.book
            worksheet = writer.sheets["Sheet1"]

            for idx, col in enumerate(results_df.columns):
                max_len = (
                    results_df[col].astype(str).map(lambda x: min(len(x), 4 * 20)).max()
                )  # Assuming 20 characters per line
                max_len = max(max_len, len(col)) + 10  # Adding some extra space
                cell_format = workbook.add_format({"text_wrap": True})

                worksheet.set_column(idx, idx, max_len, cell_format)
        except Exception as e:
            print(f"Error: {e}")
    return output_file


@main.route("/validate/upload", methods=["POST"])
def upload_file():
    if request.files["validateFile"].filename == "":
        return jsonify({"error": "ファイルが選択されていません"}), 400
    xlsx_files = []
    pdf_files = []
    for file in request.files.getlist("validateFile"):
        validate_file = file
        if validate_file.filename.lower().endswith(".xlsx"):
            xlsx_files.append(validate_file)
        elif validate_file.filename.lower().endswith(".pdf"):
            pdf_files.append(validate_file)

    default_file_name = request.form.get("defaultFile", "mojiichiran.pdf")
    base_path = getattr(sys, "_MEIPASS", os.getcwd())
    if not is_on_application():
        save_folder = os.path.join(
            base_path, os.environ.get("STANDARD_FILE_SAVE_FOLDER", "")
        )
    else:
        # save_folder = base_path
        save_folder = os.path.join(
            base_path, os.environ.get("STANDARD_FILE_SAVE_FOLDER", "")
        )
    try:
        with open(os.path.join(save_folder, default_file_name), "rb") as standard_file:
            return script_main.parallel_process_files(
                standard_file, pdf_files, xlsx_files
            )
    except FileNotFoundError:
        return (
            jsonify(
                {
                    "error": f"指定されたファイル '{default_file_name}' が見つかりませんでした。"
                }
            ),
            404,
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@main.route("/excel-compare/upload", methods=["POST"])
def upload_excel_compare_file():
    results = []
    compare_results = []
    updated_file_path = None

    if "getDataFile" in request.files:
        xlsx_compare_files = []
        for file in request.files.getlist("getDataFile"):
            if file.filename.lower().endswith(".xlsx"):
                xlsx_compare_files.append(file)
            elif "getCompareFile" in request.files:
                pass
            else:
                return (
                    jsonify({"error": "employee file is not an excel file"}),
                    400,
                )

        for file in xlsx_compare_files:
            # Save the file to a temporary location
            with NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_file:
                file.save(temp_file.name)
                temp_file_path = temp_file.name

            # Process the file
            result = script_excel_compare.get_data_from_employee(temp_file_path)
            results.append(result)

    if "getCompareFile" in request.files:
        compare_files = []
        for file in request.files.getlist("getCompareFile"):
            if file.filename.lower().endswith(".xlsx"):
                compare_files.append(file)
            elif "getDataFile" in request.files:
                pass
            else:
                return (
                    jsonify({"error": "admin file is not an excel file"}),
                    400,
                )

        for file in compare_files:
            # Save the file to a temporary location
            with NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_file:
                file.save(temp_file.name)
                temp_file_path = temp_file.name

            # Validate the file with the database
            result = script_excel_compare.validate_with_database(temp_file_path)
            compare_results.append(result)

            # Highlight incorrect user names
            workbook = load_workbook(temp_file_path)
            sheet = workbook.active
            fill = PatternFill(
                start_color="FFFF00", end_color="FFFF00", fill_type="solid"
            )
            font = Font(color="FF0000")  # Red text color

            for user in result["users_not_found_in_db"]:
                for row in sheet.iter_rows(min_row=8, max_col=6, max_row=sheet.max_row):
                    if (
                        row[4].value == user["employee_id"]
                        and row[5].value == user["name"]
                    ):
                        row[5].fill = fill  # Highlight only the "name" cell
                        row[5].font = font  # Change text color to red

            # Save the updated file to the specified directory
            save_folder = os.environ.get("VALIDATE_DOWNLOAD_SAVE_FOLDER", "")
            if not os.path.exists(save_folder):
                os.makedirs(save_folder)

            # Construct the updated file name
            original_file_name = secure_filename(file.filename)
            base_name, ext = os.path.splitext(original_file_name)
            updated_file_name = f"{base_name}_updated{ext}"
            updated_file_path = os.path.join(save_folder, updated_file_name)
            workbook.save(updated_file_path)

            # Open Finder to the directory containing the updated file
            subprocess.run(["open", save_folder], check=True)

    if not results and not compare_results:
        return jsonify({"error": "ファイルが選択されてません"}), 400

    response = {
        "data_results": results,
        "compare_results": compare_results,
    }

    if updated_file_path:
        response["updated_file"] = updated_file_path

    return jsonify(response)


@main.route("/validate/download/<filename>", methods=["GET"])
def download_file(filename):
    file_path = os.path.join(
        os.environ.get("VALIDATE_DOWNLOAD_SAVE_FOLDER", ""), filename
    )
    if not os.path.isfile(file_path):
        return jsonify({"error": "File not found"}), 404
    return send_file(file_path, max_age=200, as_attachment=True)


@main.route("/excel-compare/download/<filename>", methods=["GET"])
def download_updated_file(filename):
    file_path = os.path.join(
        os.environ.get("VALIDATE_DOWNLOAD_SAVE_FOLDER", ""), filename
    )
    if not os.path.isfile(file_path):
        return jsonify({"error": "File not found"}), 404
    return send_file(file_path, as_attachment=True)


@main.route("/test/upload", methods=["POST"])
def test_upload_file():
    xlsx_files = []
    pdf_files = []
    if len(request.files.getlist("validateFile")) == 0:
        return jsonify({"error": "Bad request"}), 400

    for file in request.files.getlist("validateFile"):
        validate_file = file
        if validate_file.filename.lower().endswith(".xlsx"):
            xlsx_files.append(validate_file.filename)
        elif validate_file.filename.lower().endswith(".pdf"):
            pdf_files.append(validate_file.filename)
    results = {"xlsx_files": xlsx_files, "pdf_files": pdf_files}
    return results


@main.route("/test/standard", methods=["GET"])
def test_standard_file():
    default_file_name = request.args.get("defaultFile")
    if default_file_name is None:
        return jsonify({"error": "Bad request"}), 400
    base_path = getattr(sys, "_MEIPASS", os.getcwd())
    if not is_on_application():
        save_folder = os.path.join(
            base_path, os.environ.get("STANDARD_FILE_SAVE_FOLDER", "")
        )
    else:
        save_folder = os.path.join(
            base_path, os.environ.get("STANDARD_FILE_SAVE_FOLDER", "")
        )
    try:
        with open(os.path.join(save_folder, default_file_name), "rb") as standard_file:
            results = {
                "file": standard_file.name,
                "file_path": os.path.join(save_folder, default_file_name),
            }
            return results
    except FileNotFoundError:
        return jsonify({"error": "File not found"}), 404
    except Exception:
        return jsonify({"error": "Internal server error"}), 500
