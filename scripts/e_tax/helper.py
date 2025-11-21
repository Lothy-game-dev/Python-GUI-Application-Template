import asyncio
from collections import Counter
import io
import multiprocessing
import os
from sys import platform
import uuid
from flask import Response, json
import pdfplumber
from openpyxl import load_workbook
from scripts.e_tax import color_xlsx_file


def get_cell_color(image, cell: tuple[int, int, int, int]):
    cropped_image = image.crop(cell)
    pixels = list(cropped_image.convert("RGB").getdata())
    color_counts = Counter(pixels)
    most_common = color_counts.most_common(1)
    return most_common[0][0]


def collect_invalid_chars_from_standard_doc(standard_file_content):

    pdf = pdfplumber.open(io.BytesIO(standard_file_content))
    invalid_chars = []

    for page in pdf.pages:
        page_image = page.to_image().original
        tables = page.find_tables()
        for table in tables:
            extracted_table = table.extract()
            for row_idx, row in enumerate(table.rows):
                for cell_idx, cell in enumerate(row.cells):
                    cell_color = get_cell_color(page_image, cell)
                    if cell_color == (192, 192, 192):
                        if extracted_table[row_idx][cell_idx] not in invalid_chars:
                            invalid_chars.append(extracted_table[row_idx][cell_idx])
    pdf.close()
    return invalid_chars


def search_pdf_file_and_get_matches(tested_pdf_file_path, text_to_search):
    pdf = pdfplumber.open(tested_pdf_file_path)
    results = []
    for page in pdf.pages:
        result_for_page = page.search(
            text_to_search,
            regex=False,
            case=False,
            main_group=0,
            return_groups=False,
            return_chars=False,
        )
        if len(result_for_page) > 0:
            for result in result_for_page:
                result["page_number"] = page.page_number
                results.append(result)
    return results


async def does_part_contain_inavalid_chars(part, collection_of_chars):
    result = {}
    for char in part:
        content = char["text"]
        if content in collection_of_chars:
            result["result"] = True
            result["firstInvalid"] = content
            return result

    result["result"] = False
    result["firstInvalid"] = ""
    return result


def read_and_divide_pdf(validate_file_content, divisor):
    pdf = pdfplumber.open(io.BytesIO(validate_file_content))
    parts_of_doc = []

    standard_length_of_each_part = max(int(len(pdf.chars) // divisor), 1)

    start_index = 0

    while True:
        parts_of_doc.append(
            pdf.chars[start_index : standard_length_of_each_part + start_index]
        )
        start_index = standard_length_of_each_part + start_index
        if start_index >= len(pdf.chars):
            break
    pdf.close()
    return parts_of_doc


def bubble_sort(arr, key):
    n = len(arr)
    for i in range(n):
        swapped = False
        for j in range(0, n - i - 1):
            if arr[j][key] > arr[j + 1][key]:
                arr[j], arr[j + 1] = arr[j + 1], arr[j]
                swapped = True
        if not swapped:
            break


def generate_unique_filename(pdf):
    # Generate a unique UUID
    unique_id = uuid.uuid4()

    # Extract the file extension
    if pdf:
        file_extension = "pdf"
    else:
        file_extension = "xlsx"

    # Construct the unique filename
    unique_filename = f"{unique_id}.{file_extension}"
    return unique_filename


async def bring_attention_to_download_folder(folder_path):
    try:
        system = platform
        match system:
            case "windows":
                await asyncio.create_subprocess_exec(
                    "explorer", "/select,", folder_path
                )
            case "darwin":
                await asyncio.create_subprocess_exec("open", "-R", folder_path)
            case "linux":
                await asyncio.create_subprocess_exec("xdg-open", folder_path)
            case _:
                print(f"Unsupported operating system: {system}")
    except Exception as e:
        print("Could not open download folder due to error: ", str(e))


def config_result_to_front_end(summary, file_handle=None, pdf=True, file_name=None):
    payload = {}
    if file_handle:
        if file_name is None:
            file_name = generate_unique_filename(pdf=pdf)

        payload["url"] = file_name

        folder_path = os.environ.get("VALIDATE_DOWNLOAD_SAVE_FOLDER", "") + file_name
        file_handle.save(folder_path)
        asyncio.run(bring_attention_to_download_folder(folder_path))
    else:
        payload = {"url": None}

    payload["result"] = {
        "type": (
            "application/pdf"
            if pdf
            else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        "summary": summary,
    }

    return Response(
        json.dumps(payload),  # JSON data
        mimetype="application/json",  # Set the MIME type to application/json
    )


def process_xlsx_file_worker(part, collected_invalid_chars, total_result_dict, order):

    result = {}
    result["invalid_cells"] = []
    result["worker_process"] = os.getpid()
    for cell in part:
        if cell.value is None:
            continue
        returned_cell = {}
        returned_cell["invalid_chars"] = []
        returned_cell["cell_coordinate"] = cell.coordinate
        returned_cell["sheet_name"] = cell.parent.title
        for char in str(cell.value):
            if char in collected_invalid_chars:
                returned_cell["invalid_chars"].append(char)

        if len(returned_cell["invalid_chars"]) > 0:
            result["invalid_cells"].append(returned_cell)
    total_result_dict[f"{order}"] = result


def parallel_processing_xlsx_file_manager(
    standard_file_content, validate_file_content, queue, divisor=2, file_name=None
):
    workbook = load_workbook(io.BytesIO(validate_file_content), rich_text=True)

    # collect all cells
    all_cells = []
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows(
            min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column
        ):
            # Iterate through all cells in the row
            for cell in row:
                all_cells.append(cell)

    # divide the cells into collections of cells
    standard_length_of_each_part = max(int(len(all_cells) // divisor), 1)
    start_index = 0
    # each part in the bellow array is a collection of cells
    parts_of_docs = []
    while True:
        parts_of_docs.append(
            all_cells[start_index : standard_length_of_each_part + start_index]
        )
        start_index = standard_length_of_each_part + start_index
        if start_index >= len(all_cells):
            break

    # collect in valid character
    collected_invalid_chars = collect_invalid_chars_from_standard_doc(
        standard_file_content
    )
    # trigger the processes to handle
    processes = []
    manager = multiprocessing.Manager()
    total_result_dict = manager.dict()
    for order, part in enumerate(parts_of_docs):
        p = multiprocessing.Process(
            target=process_xlsx_file_worker,
            args=(part, collected_invalid_chars, total_result_dict, order),
        )
        processes.append(p)
        p.start()
    # Waits for all processes to complete
    for p in processes:
        p.join()
        if p.exitcode != 0:
            raise Exception()

    total_result_dict = []
    results_to_client = []
    for order, part in enumerate(parts_of_docs):
        if len(total_result_dict[order]["invalid_cells"]) > 0:
            results_to_client.extend(total_result_dict[order]["invalid_cells"])
    if len(results_to_client) > 0:
        # color the mistakes in xlsx
        file_handle = color_xlsx_file.color_text_in_xlsx(
            workbook=workbook, validation_results=results_to_client
        )
        queue.put(
            config_result_to_front_end(
                summary=results_to_client,
                file_handle=file_handle,
                pdf=False,
                file_name=file_name,
            )
        )
    else:
        queue.put(config_result_to_front_end(summary=results_to_client))


def config_final_result_to_fe(final_response):
    payload = {}
    results = []
    urls = []

    if isinstance(final_response, list):
        parsed_data = final_response
    else:
        response_data = final_response.get_data(as_text=True)
        parsed_data = json.loads(response_data)

    for item in parsed_data:
        result = {}
        if isinstance(item, dict):
            if "result" in item:
                if isinstance(item["result"], dict):
                    if "summary" in item["result"]:
                        result["summary"] = item["result"]["summary"]
                    if "type" in item["result"]:
                        result["type"] = item["result"]["type"]
                    # print(result)
                    results.append(result)
            if "url" in item:
                urls.append(item["url"])
        else:
            print(item)

    payload["result"] = results
    payload["url"] = urls

    return Response(
        json.dumps(payload),  # JSON data
        mimetype="application/json",  # Set the MIME type to application/json
    )
